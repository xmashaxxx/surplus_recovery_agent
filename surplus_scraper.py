"""
=====================================================================
SURPLUS FUNDS SCRAPER
=====================================================================
Pulls foreclosure surplus / excess proceeds data from county sites
and saves a clean Excel pipeline file ready for skip tracing.

SUPPORTED COUNTIES (out of the box):
  - Hillsborough County, FL  (JS-rendered — uses Selenium)
  - Miami-Dade County, FL    (static HTML)
  - Fulton County, GA        (static HTML)

ADD YOUR OWN:
  Copy one of the county functions below and adapt the CSS selectors
  for your target county. Add it to COUNTY_CONFIG at the bottom.

REQUIREMENTS:
  pip install selenium openpyxl pandas requests beautifulsoup4
  Chrome + matching chromedriver (https://chromedriver.chromium.org)

USAGE:
  python surplus_scraper.py                        # runs all counties
  python surplus_scraper.py --county hillsborough  # one county only
  python surplus_scraper.py --county hillsborough --headless false
=====================================================================
"""

import argparse
import os
import re
import sys
import time
from datetime import date, datetime, timedelta

import pandas as pd
import requests
from bs4 import BeautifulSoup

# ── Selenium (only used for JS-rendered sites) ──────────────────────
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait
    SELENIUM_OK = True
except ImportError:
    SELENIUM_OK = False
    print("[WARN] Selenium not installed. JS-rendered sites will be skipped.")
    print("       Run: pip install selenium")

# ═══════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "pipeline")
DEADLINE_WARN_DAYS = 180   # flag records with < 6 months left on claim window

COUNTY_CONFIG = {
    "hillsborough": {
        "name":    "Hillsborough County, FL",
        "type":    "selenium",   # JS-rendered site
        "url":     "https://www.hillsclerk.com/court-services/foreclosure-sales",
        "state":   "FL",
        "claim_years": 1,        # FL mortgage surplus: 1 year from sale date
    },
    "miami_dade": {
        "name":    "Miami-Dade County, FL",
        "type":    "static",
        "url":     "https://www.miami-dadeclerk.com/clerkserv/surplus.asp",
        "state":   "FL",
        "claim_years": 1,
    },
    "fulton": {
        "name":    "Fulton County, GA",
        "type":    "static",
        "url":     "https://www.fultoncountyga.gov/services/property-and-real-estate/excess-funds",
        "state":   "GA",
        "claim_years": 5,        # GA: 5 years
    },
}

# ═══════════════════════════════════════════════════════════════════
#  SELENIUM DRIVER SETUP
# ═══════════════════════════════════════════════════════════════════

def get_driver(headless=True):
    """Launches Chrome. Set headless=False to watch it work."""
    if not SELENIUM_OK:
        raise RuntimeError("Selenium is not installed.")

    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )

    # Try to find chromedriver automatically; set CHROMEDRIVER_PATH env var
    # if it's in a non-standard location.
    driver_path = os.environ.get("CHROMEDRIVER_PATH", None)
    service = Service(driver_path) if driver_path else Service()
    return webdriver.Chrome(service=service, options=opts)


# ═══════════════════════════════════════════════════════════════════
#  COUNTY SCRAPERS
#  Each returns a list of dicts with these keys (use None if unknown):
#    former_owner, property_address, surplus_amount,
#    sale_date, case_number, county, state
# ═══════════════════════════════════════════════════════════════════

def scrape_hillsborough(headless=True):
    """
    Hillsborough County, FL — Selenium required (Liferay/React site).
    The foreclosure sales page shows upcoming sales; surplus amounts
    are posted as PDFs after the sale date. This scraper pulls the
    upcoming sale listings so you can check back post-sale.

    FOR SURPLUS FUNDS SPECIFICALLY: Hillsborough posts surplus fund
    notices under Circuit Civil case searches. After scraping upcoming
    sales, manually check: https://public.hillsclerk.com/CircuitCivil/
    or call (813) 276-8100 Ext. 4165 to request the surplus list.
    """
    print("  [Hillsborough] Opening browser...")
    driver = get_driver(headless=headless)
    records = []

    try:
        driver.get("https://www.hillsclerk.com/court-services/foreclosure-sales")
        wait = WebDriverWait(driver, 15)

        # Wait for page content to load
        time.sleep(3)

        # Try to find sale listings (table or list items)
        try:
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
            soup = BeautifulSoup(driver.page_source, "html.parser")
            tables = soup.find_all("table")
            for table in tables:
                rows = table.find_all("tr")[1:]  # skip header
                for row in rows:
                    cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
                    if len(cells) >= 2:
                        records.append({
                            "former_owner":      cells[1] if len(cells) > 1 else None,
                            "property_address":  cells[2] if len(cells) > 2 else None,
                            "surplus_amount":    None,  # posted post-sale
                            "sale_date":         cells[0] if len(cells) > 0 else None,
                            "case_number":       cells[3] if len(cells) > 3 else None,
                            "county":            "Hillsborough",
                            "state":             "FL",
                            "notes":             "Upcoming sale — check back post-sale for surplus",
                        })
        except Exception:
            # Fallback: grab all text content and log it
            soup = BeautifulSoup(driver.page_source, "html.parser")
            page_text = soup.get_text(separator="\n", strip=True)
            print("  [Hillsborough] Could not parse tables automatically.")
            print("  [Hillsborough] Page preview:")
            print("\n".join(page_text.split("\n")[:30]))
            print("  [Hillsborough] MANUAL STEP: Visit the page above and")
            print("  download the surplus PDF, then add records to pipeline manually.")

    finally:
        driver.quit()

    return records


def scrape_miami_dade():
    """
    Miami-Dade County, FL — static HTML.
    Surplus funds page: lists former owners and amounts directly.
    """
    print("  [Miami-Dade] Fetching page...")
    url = "https://www.miami-dadeclerk.com/clerkserv/surplus.asp"

    try:
        resp = requests.get(url, timeout=15,
                            headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
    except requests.RequestException as e:
        print(f"  [Miami-Dade] Request failed: {e}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    records = []

    # Miami-Dade typically uses a table with columns:
    # Case Number | Owner Name | Address | Sale Date | Surplus Amount
    tables = soup.find_all("table")
    for table in tables:
        headers = [th.get_text(strip=True).lower()
                   for th in table.find_all("th")]
        if not any(k in " ".join(headers) for k in ["owner", "surplus", "case"]):
            continue

        rows = table.find_all("tr")[1:]
        for row in rows:
            cells = [td.get_text(strip=True) for td in row.find_all("td")]
            if len(cells) < 3:
                continue

            # Map columns based on header order (flexible)
            col = {h: i for i, h in enumerate(headers)}
            def get(key_fragment, fallback=None):
                for h, i in col.items():
                    if key_fragment in h and i < len(cells):
                        return cells[i]
                return fallback

            records.append({
                "former_owner":      get("owner"),
                "property_address":  get("address"),
                "surplus_amount":    parse_amount(get("surplus") or get("amount")),
                "sale_date":         get("date"),
                "case_number":       get("case"),
                "county":            "Miami-Dade",
                "state":             "FL",
                "notes":             None,
            })

    if not records:
        print("  [Miami-Dade] No table data found — site structure may have changed.")
        print(f"  [Miami-Dade] Check manually: {url}")

    return records


def scrape_fulton():
    """
    Fulton County, GA — excess funds list (static page or PDF link).
    Fulton often posts a downloadable Excel/PDF. This scraper checks
    for that link and falls back to HTML parsing.
    """
    print("  [Fulton] Fetching page...")
    base_url = "https://www.fultoncountyga.gov"
    url = f"{base_url}/services/property-and-real-estate/excess-funds"
    records = []

    try:
        resp = requests.get(url, timeout=15,
                            headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
    except requests.RequestException as e:
        print(f"  [Fulton] Request failed: {e}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")

    # Check for downloadable file link first
    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = a.get_text(strip=True).lower()
        if any(ext in href.lower() for ext in [".xlsx", ".xls", ".csv", ".pdf"]):
            full_url = href if href.startswith("http") else base_url + href
            print(f"  [Fulton] Found downloadable file: {full_url}")
            print(f"  [Fulton] Download and open manually, then import to pipeline.")
            records.append({
                "former_owner":      "SEE DOWNLOAD",
                "property_address":  None,
                "surplus_amount":    None,
                "sale_date":         None,
                "case_number":       None,
                "county":            "Fulton",
                "state":             "GA",
                "notes":             f"Download list from: {full_url}",
            })
            return records

    # Fallback: parse HTML table
    tables = soup.find_all("table")
    for table in tables:
        rows = table.find_all("tr")[1:]
        for row in rows:
            cells = [td.get_text(strip=True) for td in row.find_all("td")]
            if len(cells) >= 2:
                records.append({
                    "former_owner":      cells[0] if len(cells) > 0 else None,
                    "property_address":  cells[1] if len(cells) > 1 else None,
                    "surplus_amount":    parse_amount(cells[2]) if len(cells) > 2 else None,
                    "sale_date":         cells[3] if len(cells) > 3 else None,
                    "case_number":       None,
                    "county":            "Fulton",
                    "state":             "GA",
                    "notes":             None,
                })

    if not records:
        print(f"  [Fulton] No data parsed. Visit manually: {url}")

    return records


# ═══════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════

def parse_amount(text):
    """Extract a dollar amount from messy text. Returns float or None."""
    if not text:
        return None
    text = re.sub(r"[^\d.]", "", str(text))
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(text):
    """Try multiple date formats. Returns date object or None."""
    if not text:
        return None
    for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%B %d, %Y", "%b %d, %Y"]:
        try:
            return datetime.strptime(str(text).strip(), fmt).date()
        except ValueError:
            continue
    return None


def compute_deadline(sale_date_str, claim_years):
    """Returns the claim deadline date and days remaining."""
    d = parse_date(sale_date_str)
    if not d:
        return None, None
    deadline = date(d.year + claim_years, d.month, d.day)
    days_left = (deadline - date.today()).days
    return deadline, days_left


def flag_urgency(days_left):
    if days_left is None:
        return "UNKNOWN"
    if days_left < 0:
        return "EXPIRED"
    if days_left < DEADLINE_WARN_DAYS:
        return f"URGENT — {days_left} days left"
    return f"{days_left} days left"


# ═══════════════════════════════════════════════════════════════════
#  PIPELINE BUILDER
# ═══════════════════════════════════════════════════════════════════

def build_pipeline(all_records, claim_years_map):
    rows = []
    for r in all_records:
        county_key = r.get("county", "").lower().replace("-", "_").replace(" ", "_")
        years = claim_years_map.get(county_key, 1)
        deadline, days_left = compute_deadline(r.get("sale_date"), years)

        rows.append({
            "Former Owner":       r.get("former_owner", ""),
            "Property Address":   r.get("property_address", ""),
            "County":             r.get("county", ""),
            "State":              r.get("state", ""),
            "Case Number":        r.get("case_number", ""),
            "Sale Date":          r.get("sale_date", ""),
            "Surplus Amount ($)": r.get("surplus_amount", ""),
            "Claim Deadline":     deadline,
            "Urgency":            flag_urgency(days_left),
            # Skip trace columns — fill these in manually or via API
            "Current Address":    "",
            "Phone Number":       "",
            "Email":              "",
            # Status tracking
            "Status":             "NEW",      # NEW → CONTACTED → SIGNED → FILED → PAID
            "Notes":              r.get("notes", ""),
            "Date Added":         date.today().isoformat(),
        })

    return pd.DataFrame(rows)


def save_to_excel(df):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"surplus_pipeline_{date.today().isoformat()}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Pipeline")

        ws = writer.sheets["Pipeline"]
        wb = writer.book

        # ── Column widths ──
        col_widths = {
            "A": 28, "B": 35, "C": 16, "D": 8,  "E": 18,
            "F": 14, "G": 18, "H": 16, "I": 28,
            "J": 30, "K": 18, "L": 25,
            "M": 14, "N": 40, "O": 14,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # ── Header styling ──
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_fill = PatternFill("solid", fgColor="0D3B6B")
        header_font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border

        # ── Conditional color for Urgency column (col I = index 9) ──
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            urgency_cell = row[8]  # column I (0-indexed)
            status_cell  = row[12] # column M
            val = str(urgency_cell.value or "")

            if "EXPIRED" in val:
                urgency_cell.fill = PatternFill("solid", fgColor="F7C1C1")
                urgency_cell.font = Font(color="A32D2D", bold=True, name="Arial", size=10)
            elif "URGENT" in val:
                urgency_cell.fill = PatternFill("solid", fgColor="FAC775")
                urgency_cell.font = Font(color="633806", bold=True, name="Arial", size=10)

            # Zebra rows
            for cell in row:
                if cell.row % 2 == 0:
                    if not cell.fill or cell.fill.fill_type == "none":
                        cell.fill = PatternFill("solid", fgColor="F5F7FA")
                cell.border = border
                if not cell.font or not cell.font.bold:
                    cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # ── Freeze top row ──
        ws.freeze_panes = "A2"

        # ── Instructions sheet ──
        ws_info = wb.create_sheet("Instructions")
        instructions = [
            ("SURPLUS PIPELINE TRACKER", ""),
            ("", ""),
            ("HOW TO USE THIS FILE", ""),
            ("1. Status column", "Update as you work each record: NEW → CONTACTED → SIGNED → FILED → PAID"),
            ("2. Skip Trace", "Use TLO, Spokeo, or Whitepages Pro to fill Current Address, Phone, Email"),
            ("3. Urgency column", "URGENT = less than 6 months to deadline. EXPIRED = claim window closed. Prioritize URGENT records."),
            ("4. Adding counties", "Run surplus_scraper.py again for new counties — new records append to a new dated file"),
            ("5. Surplus Amount", "If blank, the county hasn't posted the surplus figure yet. Check back post-sale."),
            ("", ""),
            ("STATUS DEFINITIONS", ""),
            ("NEW", "Just pulled from county — not yet contacted"),
            ("CONTACTED", "Reached out by phone or letter — waiting for response"),
            ("SIGNED", "Contract signed — gathering documents"),
            ("FILED", "Claim submitted to county/court — waiting for disbursement"),
            ("PAID", "Funds recovered — fee collected"),
            ("DEAD", "Claim denied, expired, or owner uninterested"),
        ]
        ws_info.column_dimensions["A"].width = 25
        ws_info.column_dimensions["B"].width = 65
        for i, (key, val) in enumerate(instructions, 1):
            cell_a = ws_info.cell(row=i, column=1, value=key)
            cell_b = ws_info.cell(row=i, column=2, value=val)
            if key and not val:
                cell_a.font = Font(bold=True, color="0D3B6B", name="Arial", size=11)
            else:
                cell_a.font = Font(bold=True, name="Arial", size=10)
                cell_b.font = Font(name="Arial", size=10)
            cell_b.alignment = Alignment(wrap_text=True)

    print(f"\n  Pipeline saved: {path}")
    return path


# ═══════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════

SCRAPERS = {
    "hillsborough": scrape_hillsborough,
    "miami_dade":   scrape_miami_dade,
    "fulton":       scrape_fulton,
}

def main():
    parser = argparse.ArgumentParser(description="Foreclosure surplus scraper")
    parser.add_argument("--county", default="all",
                        help="County to scrape: hillsborough | miami_dade | fulton | all")
    parser.add_argument("--headless", default="true",
                        help="Run Chrome headless (true/false). Use false to watch browser.")
    args = parser.parse_args()

    headless = args.headless.lower() != "false"
    target   = args.county.lower().replace("-", "_")

    if target == "all":
        targets = list(SCRAPERS.keys())
    elif target in SCRAPERS:
        targets = [target]
    else:
        print(f"Unknown county: {target}")
        print(f"Available: {', '.join(SCRAPERS.keys())}, all")
        sys.exit(1)

    all_records = []
    claim_years_map = {}

    for key in targets:
        cfg = COUNTY_CONFIG.get(key, {})
        print(f"\nScraping {cfg.get('name', key)}...")

        scraper = SCRAPERS[key]
        try:
            if key == "hillsborough" and cfg.get("type") == "selenium":
                records = scraper(headless=headless)
            else:
                records = scraper()
        except Exception as e:
            print(f"  ERROR scraping {key}: {e}")
            records = []

        county_simple = cfg.get("name", key).split(",")[0].lower().replace(" ", "_")
        claim_years_map[county_simple] = cfg.get("claim_years", 1)
        claim_years_map[key] = cfg.get("claim_years", 1)

        print(f"  Found {len(records)} records.")
        all_records.extend(records)

    if not all_records:
        print("\nNo records collected. Check county sites manually.")
        sys.exit(0)

    print(f"\nBuilding pipeline with {len(all_records)} total records...")
    df = build_pipeline(all_records, claim_years_map)

    # Show urgency summary
    urgent  = df[df["Urgency"].str.contains("URGENT",  na=False)]
    expired = df[df["Urgency"].str.contains("EXPIRED", na=False)]
    print(f"\n  Summary:")
    print(f"    Total records : {len(df)}")
    print(f"    URGENT (<6mo) : {len(urgent)}")
    print(f"    EXPIRED       : {len(expired)}")
    print(f"    Active        : {len(df) - len(urgent) - len(expired)}")

    path = save_to_excel(df)
    print(f"\nDone. Open your pipeline: {path}")


if __name__ == "__main__":
    main()
